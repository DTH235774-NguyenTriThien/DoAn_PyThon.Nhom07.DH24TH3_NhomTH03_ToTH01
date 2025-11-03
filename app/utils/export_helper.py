# app/utils/export_helper.py
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from fpdf import FPDF
from tkinter import filedialog
from app.db import fetch_query 
from tkinter import messagebox
from app.utils.utils import resource_path

def export_to_excel_from_query(parent_window, cursor, query, headers, title="Dữ liệu"):
    """
    Helper xuất dữ liệu từ query ra Excel với định dạng đẹp.
    Sử dụng filedialog để hỏi nơi lưu.
    """
    try:
        # --- Hỏi người dùng nơi lưu file ---
        now = datetime.now()
        default_filename = f"{title.replace(' ', '_')}_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        file_path = filedialog.asksaveasfilename(
            parent=parent_window,
            title=f"Lưu file Excel - {title}",
            initialfile=default_filename,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not file_path:
            return # Người dùng nhấn Hủy

        # --- Chạy truy vấn ---
        cursor.execute(query)
        rows = cursor.fetchall()

        if not rows:
            messagebox.showinfo("Không có dữ liệu", "⚠️ Không có dữ liệu để xuất!", parent=parent_window)
            return

        # --- Tạo workbook ---
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31] 

        ws.append(headers)

        for row in rows:
            ws.append([str(item).strip() if item is not None else "" for item in row])

        # --- Định dạng ---
        bold_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        for col in ws.columns:
            max_length = 0
            for cell in col:
                cell.alignment = center
                if cell.row == 1:
                    cell.font = bold_font
                    cell.fill = fill
                max_length = max(max_length, len(str(cell.value)) if cell.value else 0)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        # --- Lưu file ---
        wb.save(file_path)
        
        messagebox.showinfo("✅ Xuất thành công", f"Đã lưu file Excel:\n{file_path}", parent=parent_window)

    except Exception as e:
        messagebox.showerror("Lỗi", f"Không thể xuất file Excel: {e}", parent=parent_window)

# =========================================================
# HELPER XUẤT HÓA ĐƠN PDF
# =========================================================
class PDFReceipt(FPDF):
    """Lớp tùy chỉnh FPDF để thêm font Tiếng Việt và footer."""
    def header(self):
        pass # Header tùy chỉnh được vẽ trong hàm chính

    def footer(self):
        self.set_y(-15)
        self.set_font('DejaVu', 'I', 8)
        self.cell(0, 10, f'Trang {self.page_no()}', 0, 0, 'C')

def export_invoice_to_pdf(mahd, parent_window):
    """
    Hàm cấp cao: Lấy dữ liệu, tạo file PDF, và hỏi nơi lưu.
    """
    try:
        # 1. Lấy dữ liệu Hóa đơn (JOIN NhanVien)
        query_hd = """
            SELECT h.*, nv.HoTen AS TenNV
            FROM HoaDon h
            LEFT JOIN NhanVien nv ON h.MaNV = nv.MaNV
            WHERE h.MaHD = ?
        """
        hd_data = fetch_query(query_hd, (mahd,))
        if not hd_data: raise Exception(f"Không tìm thấy Hóa đơn {mahd}.")
        
        # 2. Lấy dữ liệu Chi tiết Hóa đơn
        cthd_data = fetch_query("""
            SELECT cthd.SoLuong, cthd.DonGia, sp.TenSP, cthd.GhiChu, cthd.ThanhTien
            FROM ChiTietHoaDon cthd
            JOIN SanPham sp ON cthd.MaSP = sp.MaSP
            WHERE cthd.MaHD = ?
        """, (mahd,))
        if not cthd_data: raise Exception(f"Không tìm thấy Chi tiết Hóa đơn cho {mahd}.")
        
        # 3. Lấy dữ liệu Khách hàng
        kh_data = None
        if hd_data[0].get('MaKH'):
            kh_result = fetch_query("SELECT TenKH FROM KhachHang WHERE MaKH = ?", (hd_data[0]['MaKH'],))
            if kh_result: kh_data = kh_result[0]
            
        hd = hd_data[0] 

        # 4. Hỏi người dùng nơi lưu file
        default_filename = f"HoaDon_{mahd}.pdf"
        file_path = filedialog.asksaveasfilename(
            parent=parent_window,
            title="Lưu hóa đơn PDF",
            initialfile=default_filename,
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")]
        )
        
        if not file_path:
            return # Người dùng nhấn Hủy

        # 5. Tạo file PDF
        pdf = PDFReceipt(orientation='P', unit='mm', format='A5') # Giấy A5
        
        # 6. Thêm Font Tiếng Việt (BẮT BUỘC)
        font_path = resource_path(os.path.join('app', 'assets', 'fonts', 'DejaVuSans.ttf'))
        
        if not os.path.exists(font_path):
                 raise Exception("Lỗi Font: Không tìm thấy file 'DejaVuSans.ttf' trong 'app/assets/fonts/'.")
                 
        pdf.add_font('DejaVu', '', font_path, uni=True)
        pdf.add_font('DejaVu', 'B', font_path, uni=True) # Bold
        pdf.add_font('DejaVu', 'I', font_path, uni=True) # Italic
        
        pdf.add_page()
        
        # --- Vẽ nội dung (HEADER) ---
        pdf.set_font('DejaVu', 'B', 16)
        pdf.cell(0, 10, "Quán Cà Phê Thiện Và Lý", 0, 1, 'C')
        pdf.set_font('DejaVu', '', 9)
        pdf.cell(0, 5, "64, Lý Thái Tổ, phường Đông Xuyên, thành phố Long Xuyên", 0, 1, 'C')
        pdf.cell(0, 5, "Hotline: 0909.123.456", 0, 1, 'C')
        pdf.ln(5) 
        
        pdf.set_font('DejaVu', 'B', 14)
        pdf.cell(0, 10, "HOA DON BAN LE", 0, 1, 'C')
        pdf.ln(5)

        # --- Vẽ thông tin Hóa đơn ---
        pdf.set_font('DejaVu', '', 10)
        now = hd.get('NgayLap', datetime.now()).strftime("%d/%m/%Y %H:%M")
        pdf.cell(0, 6, f"So HD: {mahd}         Ngay: {now}", 0, 1)
        
        thu_ngan_name = hd.get('TenNV') or hd.get('MaNV') or 'N/A'
        pdf.cell(0, 6, f"Thu ngan: {str(thu_ngan_name)}", 0, 1)
        
        if kh_data:
            pdf.cell(0, 6, f"Khach hang: {kh_data.get('TenKH', 'N/A')}", 0, 1)
        
        pdf.ln(5) 

        # --- Vẽ Bảng (Chi tiết món) ---
        pdf.set_font('DejaVu', 'B', 10)
        pdf.set_fill_color(230, 230, 230) 
        
        pdf.cell(60, 8, "Ten Mon", 1, 0, 'C', True)
        pdf.cell(15, 8, "SL", 1, 0, 'C', True)
        pdf.cell(25, 8, "Don Gia", 1, 0, 'C', True)
        pdf.cell(30, 8, "Thanh Tien", 1, 1, 'C', True)
        
        pdf.set_font('DejaVu', '', 9)
        for item in cthd_data:
            pdf.cell(60, 7, str(item.get('TenSP', 'N/A')))
            pdf.cell(15, 7, str(item.get('SoLuong', 0)), 0, 0, 'C')
            pdf.cell(25, 7, f"{int(item.get('DonGia', 0)):,}", 0, 0, 'R')
            pdf.cell(30, 7, f"{int(item.get('ThanhTien', 0)):,}", 0, 1, 'R')
            
            if item.get('GhiChu'):
                pdf.set_font('DejaVu', 'I', 8)
                pdf.set_text_color(100, 100, 100) 
                pdf.cell(60, 5, f"  (Ghi chu: {item['GhiChu']})")
                pdf.cell(70, 5, "", 0, 1) 
                pdf.set_font('DejaVu', '', 9)
                pdf.set_text_color(0, 0, 0) 

        pdf.ln(5) 

        # --- Vẽ Tổng tiền ---
        pdf.set_font('DejaVu', 'B', 10)
        tong_cong_str = f"{int(hd.get('TongTien', 0)):,} d"
        giam_gia_str = f"{int(hd.get('GiamGia', 0)):,} d"
        thanh_toan_str = f"{int(hd.get('ThanhTien', 0)):,} d"
        
        pdf.cell(100, 7, "Tong cong:", 0, 0, 'R')
        pdf.cell(30, 7, tong_cong_str, 0, 1, 'R')
        pdf.cell(100, 7, "Giam gia (diem):", 0, 0, 'R')
        pdf.cell(30, 7, giam_gia_str, 0, 1, 'R')
        
        pdf.set_font('DejaVu', 'B', 12)
        pdf.cell(100, 8, "KHACH CAN TRA:", 0, 0, 'R')
        pdf.cell(30, 8, thanh_toan_str, 0, 1, 'R')
        
        pdf.ln(10)
        pdf.set_font('DejaVu', 'I', 10)
        pdf.cell(0, 10, "Cam on quy khach! Hen gap lai!", 0, 1, 'C')

        # 7. Lưu file
        pdf.output(file_path)
        
        return True
    
    except Exception as e:
        raise e